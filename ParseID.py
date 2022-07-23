import requests
from bs4 import BeautifulSoup
import json
import openpyxl

def make_xlsx_file(spell_dict, key):
    # print(spell_dict)

    wb = openpyxl.Workbook()
    wb.remove(wb['Sheet']) # удаляем дефолтную первую страницу в таблице

    wb.create_sheet(key)
    sheet = wb[key]

    rows = 0

    for spell in spell_dict:

        rows += 1

        for i in range(1, 2+1):

            # row сдвиг по вертикали column сдвиг по горизонтали
            cell = sheet.cell(row=rows, column=i)

            if i == 1:
                cell.value = spell['name_ruru']
            if i == 2:
                cell.value = spell['spell_id']

    wb.save(f'{key}.xlsx')

x_list = []

def make_dict_with_id(script_text, key, objectName):

    item_key = 'var _ = g_items'
    class_key = 'var _ = g_classes'
    spell_key = 'var _ = g_spells'

    if key == 'items':
        text = script_text[script_text.find(item_key)+len(item_key) : script_text.find(class_key)]
    elif key == 'spells':
        text = script_text[script_text.find(spell_key)+len(spell_key) : len(script_text)]

    text = text.split(';')

    while '' in text:
        text.remove('')
    while '\n       ' in text:
        text.remove('\n       ')

    i = 0
    def name_match_check(text):
        nonlocal i

        text_0 = text[i]

        body = text_0[text_0.find(']={')+2 : len(text_0)]

        spell_dict = json.loads(body)

        if spell_dict['name_ruru'].lower() != objectName.lower():

            if i < len(text)-1:

                i += 1
                return name_match_check(text)
            else:
                print(f"{key} [{spell_dict['name_ruru']}] Что-то пошло не так.")
                return

        return text_0, spell_dict

    text_0, spell_dict = name_match_check(text)

    spell_id = text_0[text_0.find('_[')+2 : text_0.find(']={')]

    spell_dict.update({'spell_id':spell_id})

    print(f"{key} [{spell_dict['name_ruru']}] под id [{spell_dict['spell_id']}] успешно сохранён в файл.")

    x_list.append(spell_dict)

def get_html(objectName, key):
    # https://base.opiums.eu/?search=Исповедь#abilities

    url = 'https://base.opiums.eu/?search={name}#abilities'.format(name=objectName)

    source = requests.get(url)
    source_text = source.text

    soup = BeautifulSoup(source_text, 'html.parser')

    script = soup.find('body').find('script', type="text/javascript")
    # print(script)
    script_text = script.text
    # print(script_text)

    if len(script_text) == 1:
        print(f'!WARNING! {key} [{objectName}] не удалось найти в базе данных. !WARNING!')
        return

    make_dict_with_id(script_text, key, objectName)

def start_program():

    cat = input('Введите категорию для поиска: ')

    if cat != 'spells' and cat != 'items':
        print('\nВведите корректную категорию, spells либо items\n')
        start_program()
        return

    print('\n>>>>>>>>>>>>>>>>>>>> Операция началась <<<<<<<<<<<<<<<<<<<<')

    # object_list = ['Исповедь', 'Взрыв разума', 'Ментальный крик', 'Подавление боли',
    # 'Заклинание для теста без имени', 'Придание сил', 'Левитация', 'Исчадие Тьмы']
    # object_list = ['Гримуар разгневанного гладиатора', 'Перстень постепенной регенерации',
    # 'Плащ горящего заката', 'Вещь для теста без имени', 'Амулет костяного часового', 'Подвеска истинной крови']
    # object_list = ['Костечешуйный луциан']

    file = open(f'{cat}.txt', 'r', encoding='utf-8')
    object_list = file.readlines()

    for objectName in object_list:
        if '\n' in objectName:
            objectName = objectName[0:-1]
        get_html(objectName, cat)

    file.close()

    make_xlsx_file(x_list, cat)

    print('>>>>>>>>>>>>>>>>>>>> Операция завершена <<<<<<<<<<<<<<<<<<<<\n')
    input('Нажмите ENTER для закрытия программы.\n')

if __name__ == '__main__':
    print('by Friskes.\n')
    print('Приветствую, данная программа умеет собирать id у предметов или\nзаклинаний по их названию и сохранять\
 полученные данные в xlsx файл.\n\nДля работы с программой запишите интересующие вас предметы в текстовый файл\n\
с названием items.txt либо заклинания в файл spells.txt затем перезайдите в\nпрограмму и введите ниже категорию\
 для поиска items или spells соответственно.\n')
    start_program()
